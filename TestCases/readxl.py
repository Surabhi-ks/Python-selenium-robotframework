import openpyxl

wk = openpyxl.load_workbook("C:\\Users\\sks\\Desktop\\Book1.xlsx")
print("Sheets are:", wk.sheetnames)
print("Active sheet is:" , wk.active)

sh = wk['Sheet2']              #sheet level working
print("title is:", sh.title)
print(sh['A3'].value)  #way1 to get value from cell
print(sh['B3'].value)

cl = sh.cell(2,3)     #way2 to get value from cell
print("value is ", cl.value)
print(cl.row)
print(cl.column)

rows= sh.max_row
columns= sh.max_column
print("total rows are:",str(rows))
print("total columns are:", str(columns))

#way1 to fetch data
for i in range(1,rows+1):
    for j in range(1,columns+1):
        col= sh.cell(i,j)
        print("value is:", col.value)
#way2 to fetch data
for r in sh['A1':'C4']:
    for c in r:
        print(c.value)
