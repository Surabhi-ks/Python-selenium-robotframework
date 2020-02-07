import openpyxl

wk = openpyxl.load_workbook("C:\\Users\\sks\\Desktop\\Book1.xlsx")

#To know rows present in excel sheet
def fetch_number_of_rows(sheetname):
    sh = wk[sheetname]
    return sh.max_row
#To fetch data from specific cell
def fetch_cell_data(sheetname,row,cell):
    sh = wk[sheetname]
    cell= sh.cell(int(row),int(cell))
    return cell.value

print(fetch_number_of_rows("Sheet2"))
print(fetch_cell_data("Sheet2",2,1))